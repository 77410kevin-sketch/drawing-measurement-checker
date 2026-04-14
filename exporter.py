"""
匯出模組 — 將量測資料輸出為 Excel 檢表
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── 顏色定義 ────────────────────────────────────────────
HEADER_BG = "1F4E79"      # 深藍（標題列背景）
HEADER_FG = "FFFFFF"      # 白字
SUB_HEADER_BG = "2E75B6"  # 中藍（欄位標題）
ALT_ROW_BG = "D6E4F0"     # 淺藍（隔行）
FILL_AREA_BG = "FFFBCC"   # 淡黃（待填寫欄位）
BORDER_COLOR = "A9C4E2"   # 線框顏色


def _thin_border(color=BORDER_COLOR):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def export_to_excel(data: dict, output_path: str = None) -> str:
    """
    將分析結果匯出為 Excel 量測檢表

    Args:
        data: analyze_drawing_image() 回傳的 dict
        output_path: 輸出路徑（None 時自動生成）

    Returns:
        輸出檔案路徑
    """
    part_name = data.get("part_name", "Unknown")
    drawing_no = data.get("drawing_no", "N/A")
    dimensions = data.get("dimensions", [])

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = part_name.replace("/", "_").replace(" ", "_")[:30]
        output_path = f"檢表_{safe_name}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "量測檢表"

    # ── 欄寬設定 ────────────────────────────────────────
    col_widths = {
        "A": 8,   # 項次
        "B": 28,  # 量測項目
        "C": 14,  # 標稱值
        "D": 10,  # 上公差
        "E": 10,  # 下公差
        "F": 10,  # 上限值
        "G": 10,  # 下限值
        "H": 14,  # 量測結果
        "I": 10,  # 判定
        "J": 20,  # 備註
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── 標題列（R1） ────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "量 測 檢 表"
    title_cell.font = Font(name="微軟正黑體", size=18, bold=True, color=HEADER_FG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = _fill(HEADER_BG)
    title_cell.border = _thin_border("FFFFFF")

    # ── 零件資訊（R2） ───────────────────────────────────
    ws.row_dimensions[2].height = 22
    info_labels = [
        ("A2", "零件名稱"),
        ("C2", part_name),
        ("E2", "圖號"),
        ("F2", drawing_no),
        ("H2", "檢驗日期"),
        ("I2", ""),
    ]
    for cell_ref, value in info_labels:
        cell = ws[cell_ref]
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    # 標籤欄位填色
    for ref in ("A2", "E2", "H2"):
        ws[ref].font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
        ws[ref].fill = _fill(SUB_HEADER_BG)
    # 資料欄位
    for ref in ("C2", "F2", "I2"):
        ws[ref].font = Font(name="微軟正黑體", size=10)

    # 合併零件名稱和圖號的資料欄
    ws.merge_cells("C2:D2")
    ws.merge_cells("F2:G2")
    ws.merge_cells("I2:J2")

    # ── 操作人員資訊（R3） ────────────────────────────────
    ws.row_dimensions[3].height = 22
    ws.merge_cells("A3:B3")
    ws["A3"].value = "產品型號"
    ws["A3"].font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
    ws["A3"].fill = _fill(SUB_HEADER_BG)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].border = _thin_border()

    ws.merge_cells("C3:D3")
    ws["C3"].border = _thin_border()
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("E3:F3")
    ws["E3"].value = "操作人員"
    ws["E3"].font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
    ws["E3"].fill = _fill(SUB_HEADER_BG)
    ws["E3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["E3"].border = _thin_border()

    ws.merge_cells("G3:H3")
    ws["G3"].border = _thin_border()

    ws.merge_cells("I3:J3")
    ws["I3"].value = "檢驗儀器"
    ws["I3"].font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
    ws["I3"].fill = _fill(SUB_HEADER_BG)
    ws["I3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["I3"].border = _thin_border()

    # ── 欄位標題（R4） ───────────────────────────────────
    ws.row_dimensions[4].height = 24
    headers = [
        ("A4", "項次"),
        ("B4", "量測項目"),
        ("C4", "標稱值"),
        ("D4", "上公差"),
        ("E4", "下公差"),
        ("F4", "上限值"),
        ("G4", "下限值"),
        ("H4", "量測結果"),
        ("I4", "判定"),
        ("J4", "備註"),
    ]
    for cell_ref, label in headers:
        cell = ws[cell_ref]
        cell.value = label
        cell.font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
        cell.fill = _fill(SUB_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    # ── 資料列 ─────────────────────────────────────────
    start_row = 5
    for idx, dim in enumerate(dimensions):
        row = start_row + idx
        ws.row_dimensions[row].height = 20

        # 隔行填色
        row_bg = ALT_ROW_BG if idx % 2 == 0 else "FFFFFF"

        nominal = dim.get("nominal", "")
        upper_tol = dim.get("upper_tol")
        lower_tol = dim.get("lower_tol")
        unit = dim.get("unit", "mm")

        # 計算上下限
        try:
            nom_val = float(str(nominal).replace(",", ""))
            upper_limit = nom_val + float(upper_tol) if upper_tol is not None else ""
            lower_limit = nom_val + float(lower_tol) if lower_tol is not None else ""
        except (ValueError, TypeError):
            upper_limit = ""
            lower_limit = ""

        # 格式化顯示值
        def fmt(v):
            if v == "" or v is None:
                return ""
            try:
                f = float(v)
                return f"{f:g}" if f == int(f) else f"{f:.4g}"
            except (ValueError, TypeError):
                return str(v)

        nominal_display = f"{fmt(nominal)} {unit}" if nominal != "" else ""
        upper_tol_display = f"+{fmt(upper_tol)}" if upper_tol is not None else ""
        lower_tol_display = f"{fmt(lower_tol)}" if lower_tol is not None else ""
        upper_limit_display = f"{fmt(upper_limit)} {unit}" if upper_limit != "" else ""
        lower_limit_display = f"{fmt(lower_limit)} {unit}" if lower_limit != "" else ""

        row_data = [
            (f"A{row}", idx + 1),
            (f"B{row}", dim.get("name", "")),
            (f"C{row}", nominal_display),
            (f"D{row}", upper_tol_display),
            (f"E{row}", lower_tol_display),
            (f"F{row}", upper_limit_display),
            (f"G{row}", lower_limit_display),
            (f"H{row}", ""),       # 量測結果（空白待填）
            (f"I{row}", ""),       # 判定（空白待填）
            (f"J{row}", dim.get("note", "")),
        ]

        for cell_ref, value in row_data:
            cell = ws[cell_ref]
            cell.value = value
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="微軟正黑體", size=10)

            # 待填欄位（量測結果、判定）特別標色
            if cell_ref[0] in ("H", "I"):
                cell.fill = _fill(FILL_AREA_BG)
            else:
                cell.fill = _fill(row_bg)

        # 量測項目左對齊
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"J{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── 簽名列 ──────────────────────────────────────────
    sig_row = start_row + len(dimensions) + 1
    ws.row_dimensions[sig_row].height = 28
    sig_labels = [
        ("A", "B", "量測者"),
        ("C", "E", ""),
        ("F", "G", "覆核者"),
        ("H", "I", ""),
        ("J", "J", ""),
    ]
    ws.merge_cells(f"A{sig_row}:B{sig_row}")
    ws[f"A{sig_row}"].value = "量測者："
    ws[f"A{sig_row}"].font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
    ws[f"A{sig_row}"].fill = _fill(SUB_HEADER_BG)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{sig_row}"].border = _thin_border()

    ws.merge_cells(f"C{sig_row}:E{sig_row}")
    ws[f"C{sig_row}"].border = _thin_border()

    ws.merge_cells(f"F{sig_row}:G{sig_row}")
    ws[f"F{sig_row}"].value = "覆核者："
    ws[f"F{sig_row}"].font = Font(name="微軟正黑體", size=10, bold=True, color=HEADER_FG)
    ws[f"F{sig_row}"].fill = _fill(SUB_HEADER_BG)
    ws[f"F{sig_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"F{sig_row}"].border = _thin_border()

    ws.merge_cells(f"H{sig_row}:J{sig_row}")
    ws[f"H{sig_row}"].border = _thin_border()

    # ── 凍結窗格（讓標題列固定） ────────────────────────────
    ws.freeze_panes = "A5"

    # ── 列印設定 ──────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "$1:$4"

    wb.save(output_path)
    return output_path


def export_to_csv(data: dict, output_path: str = None) -> str:
    """
    將分析結果匯出為 CSV（輕量版）
    """
    import csv

    dimensions = data.get("dimensions", [])
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"checklist_{timestamp}.csv"

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["項次", "量測項目", "標稱值", "單位", "上公差", "下公差", "上限值", "下限值", "量測結果", "判定", "備註"])

        for dim in dimensions:
            nominal = dim.get("nominal", "")
            upper_tol = dim.get("upper_tol")
            lower_tol = dim.get("lower_tol")
            try:
                nom_val = float(str(nominal).replace(",", ""))
                upper_limit = nom_val + float(upper_tol) if upper_tol is not None else ""
                lower_limit = nom_val + float(lower_tol) if lower_tol is not None else ""
            except (ValueError, TypeError):
                upper_limit = ""
                lower_limit = ""

            writer.writerow([
                dim.get("item_no", ""),
                dim.get("name", ""),
                nominal,
                dim.get("unit", "mm"),
                f"+{upper_tol}" if upper_tol is not None else "",
                lower_tol if lower_tol is not None else "",
                upper_limit,
                lower_limit,
                "",  # 量測結果（空白）
                "",  # 判定（空白）
                dim.get("note", ""),
            ])

    return output_path
